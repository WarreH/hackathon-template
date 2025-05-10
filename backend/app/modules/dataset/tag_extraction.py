import pandas as pd
from datasets import load_dataset
import json
from backend.app.modules.dataset.gemini import Gemini
import openpyxl
import ast
import re
import difflib

class Data:
    def __init__(self, dataset_name="ns2agi/antwerp-osm-navigator"):
        self.dataset = load_dataset(dataset_name)
        self.train_split = self.dataset["train"]
        self.without_empty_tags = self.train_split.filter(lambda  example: example["tags"] != "{}" and example["type"] != "way")
        self.items = self.train_split.filter(lambda  example: example["tags"] != "{}" and example["type"] != "way")
        self.len = len(self.train_split)
        self.gem = Gemini()

    def filter_equals(self, dataset,filtertype, check_condition, number_of_collumns=100):
        data = dataset.filter(lambda interest: interest[filtertype] == check_condition)
        return data

    def filter_not_equals(self, dataset, filtertype, check_condition, number_of_collumns=100):
        data = dataset.filter(lambda interest: interest[filtertype] != check_condition)
        return data

    def generalize_categories(self):
        # result = data.filter_not_equals('tags', "{}")
        # items = result["tags"]
        # low_bound=1
        # bound = 500
        # length = len(result["tags"])
        # print(items[low_bound:bound])
        self.gem.prompt(
            "Do not respond to this message. I will be sending chunks (because it is to big on its own)"
            "of my dataset to you. Do not respond to these messages and just look and analyze this data."
            "The last message will start with 'last message:' you need to respond to the last message"
            "in exactly the way it describes how it wants you to respond!"
        )
        # while low_bound < length-bound-1:
        #     print("in loop")
        #     if items[low_bound:bound]:
        #         self.gem.prompt(""+str(items[low_bound:bound])+"")
        #     bound+=bound
        #     low_bound+=bound
        # print('out of loop')
        #items = self.filter_not_equals(self.train_split,'tags', "{}")["tags"]

        # items = (self.train_split
        #       .filter(lambda x: x['tags'] != {})  # drop empty
        #       .filter(lambda x: x['type'] != 'way')  # drop ways
        #       )
        #items = self.filter_not_equals(items, 'type','way')["tags"]
        #length = len(items)
        # length = 1000
        # chunk = 500
        #
        # low = 0
        # while low < length:
        #     high = min(low + chunk, length)
        #     batch = self.items[low:high]
        #
        #     if batch:  # non‐empty
        #         print(f"Sending items {low + 1}-{high}")
        #         self.gem.prompt(json.dumps(batch))
        #
        #     low = high
        # high = 500
        categories = self.gem.prompt(
            "last message: We have now send the full dataset. We want to generalize these tags, "
            "examples: bookstores and clothing stores need to be generalized to stores; Churches/castles "
            "and museums need to be generalized to cultural activities; Bar and café needs to be generalized "
            "to relaxaton. Give with this dataset the 10 most common interests/activities. Dataset:"
            ""+str(self.items[:1000]) +"Do not generate any"
            " text/explanatio and the format NEEDS to be in a list structure! like this: "
            "'[item1,item2,item3,item4,item5,item6,item7,item8,item9,item10]' with item replaced by the corresponding"
            "generalized interest/activity. This format, no duplicate categories and no other explanation needs to be "
            "followed very strictly"
            "because we are using your response in a program, so if not correct this will break the program! "
        )
        return categories

    def label_candidate(self, candidate):
        response = data.generalize_categories()
        categories = [
            "Food & Drink",
            "Cultural"
            "Shops",
            "Leisure",
            "Sports",
            "Public Transit",
            "Heritage Monuments",
            "Residential",
            "Health & Emergency",
            "Entertainement"
        ]

        resp = self.gem.prompt(
            "Given this list of categories. Look at the value of the data and choose in which categorie it"
            "needs to be! So as response give the name of the categorie that fits the value the best. So"
            "for example a categorie can be restaurant (out of the given list of categories) then a value "
            "of asian restaurant needs to be matched to this restaurant categorie. So as response you answer"
            "restaurant. If you don't have any good match with the categories, then answer with None. Do not "
            "answer any other text or explaination, just the categorie name or None. "
            "The categories: " + str(categories) + ". The value: " + candidate + "."
        )
        label = resp.text.replace("`", "").replace("python", "")

        return label

    # def _normalize(self, text: str) -> str:
    #     # strip surrounding quotes/backticks, lowercase, remove non‐alphanumerics, collapse spaces
    #     t = text.strip().strip("'\"`")
    #     t = re.sub(r'[^0-9a-zA-Z]+', ' ', t)
    #     return re.sub(r'\s+', ' ', t).strip().lower()
    #
    # def _match_category(self, raw: str, categories: li[str], cutoff: float = 0.6) -> str | None:
    #     norm_raw = self._normalize(raw)
    #     # build a list of normalized categories
    #     norm_cats = [self._normalize(c) for c in categories]
    #     # find the single best match above cutoff
    #     best = difflib.get_close_matches(norm_raw, norm_cats, n=1, cutoff=cutoff)
    #     if not best:
    #         return None
    #     idx = norm_cats.index(best[0])
    #     return categories[idx]


    def load_in_csv(self, categories,filename="catogarized_data.xlsx"):
        # 1) Create a workbook and grab the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 2) Write headers (the union of all keys in rows)
        for col_idx, header in enumerate(categories, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
        wb.save(filename)
        print(len(self.without_empty_tags))
        for data in self.without_empty_tags:
            print(data)
            resp = self.gem.prompt(
                "Given this list of categories. Look at the value of the data and choose in which categorie it"
                "needs to be! So as response give the name of the categorie that fits the value the best. So"
                "for example a categorie can be restaurant (out of the given list of categories) then a value "
                "of asian restaurant needs to be matched to this restaurant categorie. So as response you answer"
                "restaurant. If you don't have any good match with the categories, then answer with None. Do not "
                "answer any other text or explaination, just the categorie name or None. "
                "The categories: " + str(categories) + ". The value: " + str(data["tags"]) + "."
            )
            category = resp.text.replace("`", "").replace("python","")  # normalize
            try:
                print('1')
                category = eval(category)
            except Exception as e:
                print('2')
                print(f"{category} gives {e}")

            # skip empty / no-match
            if not category or category.lower() == "none":
                print("3")
                continue

            # find column index (1–10)
            try:
                print("4")
                print('categories: '+str(categories))
                for indx, cat in enumerate(categories):
                    print('cat: '+str(cat)+", categories: "+str(category))
                    if str(cat).lower() == category.lower():
                        print("index: "+str(indx))
                        col_idx = indx+1
                #col_idx = categories.index(category) + 1
            except ValueError:
                print("5")
                print(f"{category} is not in categories, skipping")
                continue

            print('6')
            # find next empty row in that column
            from openpyxl.utils import get_column_letter
            letter = get_column_letter(col_idx)
            for row in range(2, ws.max_row + 2):
                print("7")
                if ws[f"{letter}{row}"].value is None:
                    print("8")
                    ws.cell(row=row, column=col_idx, value=data["id"])
                    wb.save(filename)
                    break

        # finally, save once
        wb.save(filename)
        print("Done with the Excel file")

        # for ind, data in enumerate(self.dataset,start=2):
        #     response = self.gem.prompt(
        #         "Given this list of categories. Look at the value of the data and choose in which categorie it"
        #         "needs to be! So as response give the name of the categorie that fits the value the best. So"
        #         "for example a categorie can be restaurant (out of the given list of categories) then a value "
        #         "of asian restaurant needs to be matched to this restaurant categorie. So as response you answer"
        #         "restaurant. If you don't have any good match with the categories, then answer with None. Do not "
        #         "answer any other text or explaination, just the categorie name or None. "
        #         "The categories: "+str(categories)+". The value: "+str(data)+"."
        #     )
        #     print('gem excel: '+str(response.text))
        #     print('data: '+str(data))
        #     if response.text != "'None" and response.text != "'":
        #         try:
        #             idx = categories.index(response.text)
        #             column = chr(ord('A') + idx)
        #             next_row = ws.max_row + 1
        #             for row in range(1, ws.max_row + 2):  # +2 so we check one past the current max
        #                 cell = ws[f"{column}{row}"]
        #                 if cell.value is None:
        #                     next_row = row
        #                     break
        #             cell = ws.cell(row=next_row, column=column, value=data["id"])
        #             wb.save(filename)
        #         except ValueError:
        #             print(f"'{response.text}' is not in the list")
        # print('Done with the excell file')


    def create_list_from_gem(self, gem_response):
        categories = gem_response.text.split("[")[1].split("]")[0]
        list = ast.literal_eval(categories)
        return list

if __name__ == "__main__":
    data = Data()
    response = data.generalize_categories()
    print('response gem: '+response.text)
    list = data.create_list_from_gem(response)
    data.load_in_csv(list)
    print("\nScript finished.")
